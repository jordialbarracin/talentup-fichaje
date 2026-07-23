"""
TalentUP Fichaje — Reports router.
GET /api/reports/hours, GET /api/reports/incidents, GET /api/reports/export
GET /api/reports/inspection, GET /api/reports/absenteeism, GET /api/reports/labor-costs
"""
import io
import os
import uuid
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from math import ceil
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.clock_in import ClockIn
from app.models.employee import Employee
from app.models.incident import Incident
from app.models.schedule import Schedule
from app.models.shift import Shift
from app.models.user import User
from app.auth import require_manager, get_current_user
from app.pagination import paginate

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _parse_date(date_str: str, field_name: str) -> date:
    """Parse ISO date string, raise 400 on invalid format."""
    try:
        return date.fromisoformat(date_str)
    except (ValueError, TypeError):
        raise HTTPException(
            status_code=400,
            detail=f"Formato de fecha inválido para '{field_name}': {date_str}. Use YYYY-MM-DD.",
        )


def _resolve_tenant_id(current_user: User, query_tenant_id: Optional[str] = None):
    """Resolve tenant_id: super_admin can pass optional tenant_id, others use their own."""
    if current_user.role == "super_admin":
        return query_tenant_id  # None means all tenants
    return current_user.tenant_id


@router.get("/hours")
async def report_hours(
    employee_id: Optional[str] = None,
    date_from: str = Query(...),
    date_to: str = Query(...),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    tenant_id: Optional[str] = Query(None, description="Solo para super_admin: filtrar por tenant"),
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db),
):
    """Hours worked per employee in a date range."""
    tid = _resolve_tenant_id(current_user, tenant_id)
    start_date = _parse_date(date_from, "date_from")
    end_date = _parse_date(date_to, "date_to")

    # Resolve which employee IDs to include and paginate on employee level
    if tid:
        emp_query = select(Employee.id, Employee.name).where(Employee.tenant_id == tid)
    else:
        emp_query = select(Employee.id, Employee.name)
    if employee_id:
        emp_query = emp_query.where(Employee.id == employee_id)
    emp_query = emp_query.order_by(Employee.name)

    # Count total for pagination metadata
    count_q = select(func.count()).select_from(emp_query.subquery())
    total = (await db.execute(count_q)).scalar() or 0
    # Manual pagination for multi-column select (paginate() uses .scalars() which flattens tuples)
    offset = (page - 1) * limit
    emp_rows = (await db.execute(emp_query.offset(offset).limit(limit))).all()
    emp_ids = [row[0] for row in emp_rows]
    emp_names = {row[0]: row[1] for row in emp_rows}
    pages = int(ceil(total / limit)) if limit else 1

    day_start = datetime.combine(start_date, time.min, tzinfo=timezone.utc)
    day_end = datetime.combine(end_date, time.max, tzinfo=timezone.utc)

    # Aggregate daily totals in the DB using a CTE pairing consecutive in/out rows.
    clock_pairs_cte = (
        select(
            ClockIn.employee_id,
            func.date(ClockIn.timestamp).label("clock_date"),
            ClockIn.type.label("type"),
            func.strftime("%s", ClockIn.timestamp).label("in_epoch"),
            func.strftime("%s", func.lead(ClockIn.timestamp).over(
                partition_by=ClockIn.employee_id,
                order_by=ClockIn.timestamp,
            )).label("out_epoch"),
            func.lead(ClockIn.type).over(
                partition_by=ClockIn.employee_id,
                order_by=ClockIn.timestamp,
            ).label("next_type"),
        )
        .where(
            ClockIn.timestamp >= day_start,
            ClockIn.timestamp <= day_end,
            ClockIn.is_cancelled == False,
            ClockIn.type.in_(["in", "out"]),
        )
    )
    if tid:
        clock_pairs_cte = clock_pairs_cte.where(ClockIn.tenant_id == tid)
    if employee_id:
        clock_pairs_cte = clock_pairs_cte.where(ClockIn.employee_id == employee_id)
    else:
        if emp_ids:
            clock_pairs_cte = clock_pairs_cte.where(ClockIn.employee_id.in_(emp_ids))
        else:
            clock_pairs_cte = clock_pairs_cte.where(False)
    clock_pairs_cte = clock_pairs_cte.cte("clock_pairs")

    daily_totals = select(
        clock_pairs_cte.c.employee_id,
        clock_pairs_cte.c.clock_date,
        func.round(func.sum(clock_pairs_cte.c.out_epoch - clock_pairs_cte.c.in_epoch) / 3600.0, 2).label("daily_hours"),
        func.sum(clock_pairs_cte.c.out_epoch - clock_pairs_cte.c.in_epoch).label("daily_seconds"),
    ).where(
        clock_pairs_cte.c.type == "in",
        clock_pairs_cte.c.next_type == "out",
        clock_pairs_cte.c.out_epoch.is_not(None),
    ).group_by(
        clock_pairs_cte.c.employee_id,
        clock_pairs_cte.c.clock_date,
    )

    result = await db.execute(daily_totals)
    daily_rows = result.all()

    totals = {}
    daily_breakdown = defaultdict(lambda: defaultdict(float))
    for emp_id, clock_date, hours, seconds in daily_rows:
        totals[emp_id] = totals.get(emp_id, 0) + (seconds or 0)
        # clock_date comes as string from SQLite func.date(); normalize to ISO string
        date_key = clock_date.isoformat() if hasattr(clock_date, 'isoformat') else str(clock_date)
        daily_breakdown[emp_id][date_key] += float(hours or 0)

    report = []
    for emp_id in emp_ids:
        total_seconds = totals.get(emp_id, 0)
        report.append({
            "employee_id": str(emp_id),
            "employee_name": emp_names.get(emp_id, "Desconocido"),
            "total_hours": round(total_seconds / 3600, 2),
            "total_minutes": int(total_seconds / 60),
            "days": len(daily_breakdown[emp_id]),
            "daily_hours": dict(daily_breakdown[emp_id]),
        })

    return {
        "date_from": date_from,
        "date_to": date_to,
        "tenant_id": str(tid) if tid else "all",
        "employees": report,
        "page": page,
        "limit": limit,
        "total": total,
        "pages": pages,
    }


@router.get("/incidents")
async def report_incidents(
    employee_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    incident_type: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    tenant_id: Optional[str] = Query(None, description="Solo para super_admin: filtrar por tenant"),
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db),
):
    """List incidents with filters."""
    # Aggregate incident counts by type instead of loading all rows into memory.
    tid = _resolve_tenant_id(current_user, tenant_id)
    query = select(
        Incident.incident_type,
        func.count(Incident.id).label("count"),
    )
    if tid:
        query = query.where(Incident.tenant_id == tid)
    if employee_id:
        query = query.where(Incident.employee_id == employee_id)
    if date_from:
        query = query.where(Incident.date >= _parse_date(date_from, "date_from"))
    if date_to:
        query = query.where(Incident.date <= _parse_date(date_to, "date_to"))
    if incident_type:
        query = query.where(Incident.incident_type == incident_type)
    query = query.group_by(Incident.incident_type).order_by(func.count(Incident.id).desc())
    agg_result = await db.execute(query)
    summary = [{"incident_type": t, "count": c} for t, c in agg_result.all()]

    # Paginate raw incidents, then enrich page items only.
    query = select(Incident)
    if tid:
        query = query.where(Incident.tenant_id == tid)
    if employee_id:
        query = query.where(Incident.employee_id == employee_id)
    if date_from:
        query = query.where(Incident.date >= _parse_date(date_from, "date_from"))
    if date_to:
        query = query.where(Incident.date <= _parse_date(date_to, "date_to"))
    if incident_type:
        query = query.where(Incident.incident_type == incident_type)
    query = query.order_by(Incident.date.desc(), Incident.created_at.desc())

    page_result = await paginate(db, query, page, limit, item_transform=lambda i: i)
    incidents = page_result["items"]

    # Enrich with employee names for the page only.
    emp_ids = {i.employee_id for i in incidents}
    emp_map = {}
    if emp_ids:
        emp_result = await db.execute(
            select(Employee.id, Employee.name).where(Employee.id.in_(emp_ids))
        )
        emp_map = {e_id: name for e_id, name in emp_result.all()}

    items = []
    for inc in incidents:
        item = inc.to_dict()
        item["employee_name"] = emp_map.get(inc.employee_id, "Desconocido")
        items.append(item)

    page_result["items"] = items
    page_result["summary"] = summary
    return page_result


async def _build_export_data(
    db: AsyncSession,
    *,
    tid: Optional[str],
    date_from: str,
    date_to: str,
    employee_id: Optional[str] = None,
    page: int = 1,
    limit: int = 100,
):
    """Gather paginated employee data for report export."""
    start_date = _parse_date(date_from, "date_from")
    end_date = _parse_date(date_to, "date_to")

    # Paginate employees and only load relevant fields.
    if tid:
        emp_query = select(Employee.id, Employee.name, Employee.dni).where(Employee.tenant_id == tid)
    else:
        emp_query = select(Employee.id, Employee.name, Employee.dni)
    if employee_id:
        emp_query = emp_query.where(Employee.id == employee_id)
    emp_query = emp_query.order_by(Employee.name)
    export_offset = (page - 1) * limit
    emp_rows = (await db.execute(emp_query.offset(export_offset).limit(limit))).all()
    emp_ids = [row[0] for row in emp_rows]
    emp_info = {row[0]: (row[1], row[2] or "") for row in emp_rows}

    day_start = datetime.combine(start_date, time.min, tzinfo=timezone.utc)
    day_end = datetime.combine(end_date, time.max, tzinfo=timezone.utc)

    # Aggregate clock entries at DB level using window functions.
    clock_pairs_cte = (
        select(
            ClockIn.employee_id,
            func.date(ClockIn.timestamp).label("clock_date"),
            ClockIn.type.label("type"),
            func.strftime("%s", ClockIn.timestamp).label("in_epoch"),
            func.strftime("%s", func.lead(ClockIn.timestamp).over(
                partition_by=ClockIn.employee_id,
                order_by=ClockIn.timestamp,
            )).label("out_epoch"),
            func.lead(ClockIn.type).over(
                partition_by=ClockIn.employee_id,
                order_by=ClockIn.timestamp,
            ).label("next_type"),
        )
        .where(
            ClockIn.timestamp >= day_start,
            ClockIn.timestamp <= day_end,
            ClockIn.is_cancelled == False,
            ClockIn.type.in_(["in", "out"]),
        )
    )
    if tid:
        clock_pairs_cte = clock_pairs_cte.where(ClockIn.tenant_id == tid)
    if employee_id:
        clock_pairs_cte = clock_pairs_cte.where(ClockIn.employee_id == employee_id)
    elif emp_ids:
        clock_pairs_cte = clock_pairs_cte.where(ClockIn.employee_id.in_(emp_ids))
    else:
        clock_pairs_cte = clock_pairs_cte.where(False)
    clock_pairs_cte = clock_pairs_cte.cte("clock_pairs")

    entries_query = (
        select(
            clock_pairs_cte.c.employee_id,
            clock_pairs_cte.c.clock_date,
            func.min(clock_pairs_cte.c.in_epoch).label("first_in_epoch"),
            func.max(clock_pairs_cte.c.out_epoch).label("last_out_epoch"),
            func.sum(clock_pairs_cte.c.out_epoch - clock_pairs_cte.c.in_epoch).label("total_seconds"),
        )
        .where(
            clock_pairs_cte.c.type == "in",
            clock_pairs_cte.c.next_type == "out",
            clock_pairs_cte.c.out_epoch.is_not(None),
        )
        .group_by(clock_pairs_cte.c.employee_id, clock_pairs_cte.c.clock_date)
        .order_by(clock_pairs_cte.c.employee_id, clock_pairs_cte.c.clock_date)
    )
    result = await db.execute(entries_query)
    rows = result.all()

    entries_by_emp = defaultdict(list)
    totals_by_emp = defaultdict(int)
    for emp_id, clock_date, first_in_epoch, last_out_epoch, total_seconds in rows:
        if not first_in_epoch or not last_out_epoch:
            continue
        try:
            first_in_epoch_int = int(first_in_epoch)
        except (TypeError, ValueError):
            continue
        try:
            last_out_epoch_int = int(last_out_epoch)
        except (TypeError, ValueError):
            continue
        first_in_dt = datetime.fromtimestamp(first_in_epoch_int, tz=timezone.utc)
        last_out_dt = datetime.fromtimestamp(last_out_epoch_int, tz=timezone.utc)
        entries_by_emp[emp_id].append({
            "date": clock_date.isoformat() if hasattr(clock_date, 'isoformat') else str(clock_date),
            "in": first_in_dt.strftime("%H:%M"),
            "out": last_out_dt.strftime("%H:%M"),
            "hours": round(total_seconds / 3600, 2),
        })
        totals_by_emp[emp_id] += total_seconds or 0

    report_data = []
    for emp_id in emp_ids:
        name, dni = emp_info.get(emp_id, ("Desconocido", ""))
        total_seconds = totals_by_emp.get(emp_id, 0)
        report_data.append({
            "name": name,
            "dni": dni,
            "total_hours": round(total_seconds / 3600, 2),
            "entries": entries_by_emp.get(emp_id, []),
        })

    return report_data, tid


@router.get("/export")
async def export_report(
    format: str = Query("pdf", pattern="^(pdf|excel)$"),
    date_from: str = Query(...),
    date_to: str = Query(...),
    employee_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(100, ge=1, le=500),
    tenant_id: Optional[str] = Query(None, description="Solo para super_admin: filtrar por tenant"),
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db),
):
    """Export hours report as PDF or Excel binary file."""
    report_data, tid = await _build_export_data(
        db,
        tid=_resolve_tenant_id(current_user, tenant_id),
        date_from=date_from,
        date_to=date_to,
        employee_id=employee_id,
        page=page,
        limit=limit,
    )

    if format == "pdf":
        return _generate_pdf(report_data, date_from, date_to)
    return _generate_excel(report_data, date_from, date_to)


# Keep the synchronous generators available for in-process / backwards-compatible use.

def _generate_pdf(data, date_from, date_to):
    """Generate PDF report using reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        title=f"Informe de Fichajes {date_from} a {date_to}",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=18,
        textColor=colors.HexColor("#FF6B35"),
        spaceAfter=20,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.gray,
        spaceAfter=20,
    )

    elements = []
    elements.append(Paragraph("TalentUP Fichaje — Informe de Registro", title_style))
    elements.append(Paragraph(f"Período: {date_from} a {date_to}", subtitle_style))
    elements.append(Spacer(1, 10))

    for emp in data:
        elements.append(Paragraph(
            f"<b>{emp['name']}</b>  |  DNI: {emp['dni']}  |  Total: {emp['total_hours']}h",
            styles["Heading2"],
        ))
        elements.append(Spacer(1, 5))

        if not emp["entries"]:
            elements.append(Paragraph("Sin fichajes en este período", styles["Normal"]))
        else:
            table_data = [["Fecha", "Entrada", "Salida", "Horas"]]
            for entry in emp["entries"]:
                table_data.append([
                    entry["date"],
                    entry["in"],
                    entry["out"],
                    f"{entry['hours']}h",
                ])
            # Total row
            table_data.append(["", "", "TOTAL:", f"{emp['total_hours']}h"])

            col_widths = [100, 80, 80, 80]
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF6B35")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -2), 0.5, colors.grey),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF3EB")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F8F9FA")]),
            ]))
            elements.append(table)

        elements.append(Spacer(1, 15))

    # Footer with legal notice
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        "Documento generado por TalentUP Fichaje. "
        "Registro de jornada laboral conforme al RD-ley 8/2019 art. 34.9 ET. "
        "Conservación mínima: 4 años.",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7, textColor=colors.grey),
    ))

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=fichajes_{date_from}_{date_to}.pdf",
        },
    )


def _generate_excel(data, date_from, date_to):
    """Generate Excel report using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Fichajes"

    # Header
    ws.merge_cells("A1:D1")
    ws["A1"] = f"TalentUP Fichaje — Informe {date_from} a {date_to}"
    ws["A1"].font = Font(size=14, bold=True, color="FF6B35")
    ws["A1"].alignment = Alignment(horizontal="center")

    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row = 3
    for emp in data:
        # Employee header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=f"{emp['name']} — Total: {emp['total_hours']}h")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        # Table header
        for col, h in enumerate(["Fecha", "Entrada", "Salida", "Horas"], 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        row += 1

        # Data rows
        for entry in emp["entries"]:
            ws.cell(row=row, column=1, value=entry["date"]).border = thin_border
            ws.cell(row=row, column=2, value=entry["in"]).border = thin_border
            ws.cell(row=row, column=3, value=entry["out"]).border = thin_border
            ws.cell(row=row, column=4, value=entry["hours"]).border = thin_border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
            row += 1

        # Total row
        ws.cell(row=row, column=3, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4, value=emp["total_hours"]).font = Font(bold=True)
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
        row += 2  # blank row between employees

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=fichajes_{date_from}_{date_to}.xlsx",
        },
    )


@router.get("/inspection")
async def report_inspection(
    date_from: str = Query(...),
    date_to: str = Query(...),
    employee_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    tenant_id: Optional[str] = Query(None, description="Solo para super_admin"),
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db),
):
    """
    Informe para Inspección de Trabajo (RD-ley 8/2019).
    Incluye: registro diario de fichajes, horas extras, incidencias, anulaciones.
    """
    tid = _resolve_tenant_id(current_user, tenant_id)
    start_date = _parse_date(date_from, "date_from")
    end_date = _parse_date(date_to, "date_to")

    # Get employees
    if tid:
        emp_query = select(Employee).where(Employee.tenant_id == tid)
    else:
        emp_query = select(Employee)
    if employee_id:
        emp_query = emp_query.where(Employee.id == employee_id)
    emp_query = emp_query.order_by(Employee.name)
    page_result = await paginate(db, emp_query, page, limit, item_transform=lambda e: e)
    employees = page_result["items"]
    total_employees = page_result["total"]

    # Get employee IDs for targeted dependent queries.
    emp_ids = [emp.id for emp in employees]

    # Get tenant info
    tenant_info = {}
    if tid:
        from app.models.tenant import Tenant
        t_result = await db.execute(select(Tenant).where(Tenant.id == tid))
        t = t_result.scalar_one_or_none()
        if t:
            tenant_info = {"name": t.name, "legal_name": t.legal_name, "cif": t.cif, "address": t.address}

    # Get clock-ins
    day_start = datetime.combine(start_date, time.min, tzinfo=timezone.utc)
    day_end = datetime.combine(end_date, time.max, tzinfo=timezone.utc)
    clock_query = select(ClockIn).where(
        ClockIn.timestamp >= day_start,
        ClockIn.timestamp <= day_end,
    )
    if tid:
        clock_query = clock_query.where(ClockIn.tenant_id == tid)
    if emp_ids:
        clock_query = clock_query.where(ClockIn.employee_id.in_(emp_ids))
    clock_query = clock_query.order_by(ClockIn.employee_id, ClockIn.timestamp).limit(500)
    result = await db.execute(clock_query)
    all_clock_ins = result.scalars().all()

    # Get incidents
    inc_query = select(Incident).where(
        Incident.date >= start_date,
        Incident.date <= end_date,
    )
    if tid:
        inc_query = inc_query.where(Incident.tenant_id == tid)
    if emp_ids:
        inc_query = inc_query.where(Incident.employee_id.in_(emp_ids))
    inc_query = inc_query.order_by(Incident.employee_id, Incident.date).limit(500)
    result = await db.execute(inc_query)
    all_incidents = result.scalars().all()

    # Count total incidents across the full period (not just this page) for summary.
    total_incidents_count_query = select(func.count()).select_from(Incident).where(
        Incident.date >= start_date,
        Incident.date <= end_date,
    )
    if tid:
        total_incidents_count_query = total_incidents_count_query.where(Incident.tenant_id == tid)
    total_incidents_result = await db.execute(total_incidents_count_query)
    total_incidents = total_incidents_result.scalar() or 0

    from collections import defaultdict
    clock_by_emp = defaultdict(list)
    for ci in all_clock_ins:
        clock_by_emp[ci.employee_id].append(ci)

    inc_by_emp = defaultdict(list)
    for inc in all_incidents:
        inc_by_emp[inc.employee_id].append(inc)

    report_employees = []
    for emp in employees:
        emp_clock = clock_by_emp.get(emp.id, [])
        emp_inc = inc_by_emp.get(emp.id, [])

        # Daily breakdown
        daily_clock = defaultdict(list)
        for ci in emp_clock:
            daily_clock[ci.timestamp.date()].append(ci)

        daily_entries = []
        total_seconds = 0
        for day_date, day_clock in sorted(daily_clock.items()):
            first_in = None
            last_out = None
            breaks = []
            for ci in day_clock:
                if ci.type == "in":
                    if first_in is None or ci.timestamp < first_in:
                        first_in = ci.timestamp
                elif ci.type == "out":
                    if last_out is None or ci.timestamp > last_out:
                        last_out = ci.timestamp
                elif ci.type == "break_start":
                    breaks.append({"start": ci.timestamp, "end": None})
                elif ci.type == "break_end" and breaks and breaks[-1]["end"] is None:
                    breaks[-1]["end"] = ci.timestamp

            if first_in and last_out:
                delta = (last_out - first_in).total_seconds()
                total_seconds += delta
                daily_entries.append({
                    "date": day_date.isoformat(),
                    "entry": first_in.strftime("%H:%M"),
                    "exit": last_out.strftime("%H:%M"),
                    "hours": round(delta / 3600, 2),
                    "breaks": [{"start": b["start"].strftime("%H:%M") if b["start"] else None,
                                "end": b["end"].strftime("%H:%M") if b["end"] else None} for b in breaks],
                })

        report_employees.append({
            "employee_id": str(emp.id),
            "name": emp.name,
            "dni": emp.dni or "",
            "categoria": emp.categoria_profesional or "",
            "total_hours": round(total_seconds / 3600, 2),
            "daily_entries": daily_entries,
            "incidents": [i.to_dict() for i in emp_inc],
        })

    return {
        "tenant": tenant_info,
        "period": {"from": date_from, "to": date_to},
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "legal_notice": "Documento generado por TalentUP Fichaje. "
                        "Registro de jornada laboral conforme al RD-ley 8/2019 art. 34.9 ET. "
                        "Conservación mínima: 4 años.",
        "employees": report_employees,
        "page": page,
        "limit": limit,
        "total_employees": total_employees,
        "summary": {
            "total_employees": total_employees,
            "total_incidents": total_incidents,
        }
    }


@router.get("/absenteeism")
async def report_absenteeism(
    date_from: str = Query(...),
    date_to: str = Query(...),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    tenant_id: Optional[str] = Query(None, description="Solo para super_admin"),
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db),
):
    """
    Informe de absentismo: tasa global, desglose por tipo, por empleado.
    """
    tid = _resolve_tenant_id(current_user, tenant_id)
    start_date = _parse_date(date_from, "date_from")
    end_date = _parse_date(date_to, "date_to")

    if tid:
        emp_query = select(Employee).where(Employee.tenant_id == tid)
    else:
        emp_query = select(Employee)
    emp_query = emp_query.order_by(Employee.name)
    page_result = await paginate(db, emp_query, page, limit, item_transform=lambda e: e)
    employees = page_result["items"]
    total_employees = page_result["total"]
    emp_ids = [emp.id for emp in employees]

    # Get vacation requests in range
    from app.models.vacation_request import VacationRequest
    vac_query = select(VacationRequest).where(
        VacationRequest.start_date >= start_date,
        VacationRequest.end_date <= end_date,
        VacationRequest.status == "approved",
    )
    if tid:
        vac_query = vac_query.where(VacationRequest.tenant_id == tid)
    if emp_ids:
        vac_query = vac_query.where(VacationRequest.employee_id.in_(emp_ids))
    vac_query = vac_query.order_by(VacationRequest.employee_id).limit(500)
    result = await db.execute(vac_query)
    vacations = result.scalars().all()

    # Get leave in range
    from app.models.leave import Leave
    leave_query = select(Leave).where(
        Leave.start_date >= start_date,
        Leave.end_date <= end_date,
    )
    if tid:
        leave_query = leave_query.where(Leave.tenant_id == tid)
    if emp_ids:
        leave_query = leave_query.where(Leave.employee_id.in_(emp_ids))
    leave_query = leave_query.order_by(Leave.employee_id).limit(500)
    result = await db.execute(leave_query)
    leaves = result.scalars().all()

    # Get no_show incidents
    inc_query = select(Incident).where(
        Incident.date >= start_date,
        Incident.date <= end_date,
        Incident.incident_type.in_(["no_clock_in", "no_show", "ausencia_no_justificada"]),
    )
    if tid:
        inc_query = inc_query.where(Incident.tenant_id == tid)
    if emp_ids:
        inc_query = inc_query.where(Incident.employee_id.in_(emp_ids))
    inc_query = inc_query.order_by(Incident.employee_id).limit(500)
    result = await db.execute(inc_query)
    no_shows = result.scalars().all()

    # Global totals must still reflect all employees, not just this page.
    total_days = (end_date - start_date).days + 1
    total_possible_days = total_employees * total_days

    # Paginate employee breakdown after computing rates/sorting.
    from collections import defaultdict
    emp_vac = defaultdict(list)
    for v in vacations:
        emp_vac[v.employee_id].append(v)
    emp_leave = defaultdict(list)
    for l in leaves:
        emp_leave[l.employee_id].append(l)
    emp_noshow = defaultdict(list)
    for n in no_shows:
        emp_noshow[n.employee_id].append(n)

    employee_breakdown = []
    for emp in employees:
        vac_days = sum(int(v.total_days or 0) for v in emp_vac.get(emp.id, []))
        leave_days = sum(l.total_days or 0 for l in emp_leave.get(emp.id, []))
        noshow_days = len(emp_noshow.get(emp.id, []))
        total_emp_absence = vac_days + leave_days + noshow_days
        emp_rate = round((total_emp_absence / total_days * 100), 2) if total_days > 0 else 0

        employee_breakdown.append({
            "employee_id": str(emp.id),
            "name": emp.name,
            "vacation_days": vac_days,
            "leave_days": leave_days,
            "no_show_days": noshow_days,
            "total_absence_days": total_emp_absence,
            "absenteeism_rate": emp_rate,
        })

    # Sort by rate desc
    employee_breakdown.sort(key=lambda x: x["absenteeism_rate"], reverse=True)

    # Global aggregates: derive from full-tenant data so pagination doesn't distort rates.
    global_vac_query = select(func.sum(VacationRequest.total_days)).where(
        VacationRequest.start_date >= start_date,
        VacationRequest.end_date <= end_date,
        VacationRequest.status == "approved",
    )
    global_leave_query = select(func.sum(Leave.total_days)).where(
        Leave.start_date >= start_date,
        Leave.end_date <= end_date,
    )
    global_noshow_query = select(func.count()).select_from(Incident).where(
        Incident.date >= start_date,
        Incident.date <= end_date,
        Incident.incident_type.in_(["no_clock_in", "no_show", "ausencia_no_justificada"]),
    )
    if tid:
        global_vac_query = global_vac_query.where(VacationRequest.tenant_id == tid)
        global_leave_query = global_leave_query.where(Leave.tenant_id == tid)
        global_noshow_query = global_noshow_query.where(Incident.tenant_id == tid)
    global_vac_sum = (await db.execute(global_vac_query)).scalar() or 0
    global_leave_sum = (await db.execute(global_leave_query)).scalar() or 0
    global_noshow_count = (await db.execute(global_noshow_query)).scalar() or 0

    total_absence_days = int(global_vac_sum) + global_leave_sum + global_noshow_count
    absenteeism_rate = round((total_absence_days / total_possible_days * 100), 2) if total_possible_days > 0 else 0

    # Slice top 5 from paginated breakdown.
    top_5_absentees = employee_breakdown[:5]

    return {
        "period": {"from": date_from, "to": date_to},
        "global_absenteeism_rate": absenteeism_rate,
        "total_absence_days": total_absence_days,
        "total_possible_days": total_possible_days,
        "breakdown": {
            "vacations": int(global_vac_sum),
            "leave": global_leave_sum,
            "no_show": global_noshow_count,
        },
        "employees": employee_breakdown,
        "top_5_absentees": top_5_absentees,
        "page": page,
        "limit": limit,
        "total_employees": total_employees,
    }


@router.get("/labor-costs")
async def report_labor_costs(
    month: int = Query(...),
    year: int = Query(...),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    tenant_id: Optional[str] = Query(None, description="Solo para super_admin"),
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db),
):
    """
    Informe de costes laborales: salarios, complementos, SS, IRPF.
    """
    tid = _resolve_tenant_id(current_user, tenant_id)

    from app.models.payroll import Payroll
    query = select(Payroll).where(
        Payroll.year == year,
        Payroll.month == month,
    )
    if tid:
        query = query.where(Payroll.tenant_id == tid)
    query = query.order_by(Payroll.employee_id)
    page_result = await paginate(db, query, page, limit, item_transform=lambda p: p)
    payrolls = page_result["items"]
    total_payrolls = page_result["total"]

    # Get employees for names
    emp_ids = {p.employee_id for p in payrolls}
    emp_result = await db.execute(select(Employee).where(Employee.id.in_(emp_ids)))
    emp_map = {e.id: e for e in emp_result.scalars().all()}

    total_base = 0
    total_night = 0
    total_holiday = 0
    total_seniority = 0
    total_overtime = 0
    total_ss = 0
    total_irpf = 0
    total_gross = 0
    total_net = 0

    employee_costs = []
    for p in payrolls:
        emp = emp_map.get(p.employee_id)
        total_base += float(p.base_salary or 0)
        total_night += float(p.night_plus or 0)
        total_holiday += float(p.holiday_plus or 0)
        total_seniority += float(p.seniority_plus or 0)
        total_overtime += float(p.overtime_amount or 0)
        total_ss += float(p.ss_deduction or 0)
        total_irpf += float(p.irpf_deduction or 0)
        total_gross += float(p.gross_total or 0)
        total_net += float(p.net_total or 0)

        employee_costs.append({
            "employee_id": str(p.employee_id),
            "name": emp.name if emp else "Desconocido",
            "category": emp.categoria_profesional if emp else "",
            "base_salary": float(p.base_salary or 0),
            "night_plus": float(p.night_plus or 0),
            "holiday_plus": float(p.holiday_plus or 0),
            "seniority_plus": float(p.seniority_plus or 0),
            "overtime_amount": float(p.overtime_amount or 0),
            "gross_total": float(p.gross_total or 0),
            "ss_deduction": float(p.ss_deduction or 0),
            "irpf_deduction": float(p.irpf_deduction or 0),
            "net_total": float(p.net_total or 0),
        })

    return {
        "period": f"{month}/{year}",
        "summary": {
            "total_base_salary": round(total_base, 2),
            "total_night_plus": round(total_night, 2),
            "total_holiday_plus": round(total_holiday, 2),
            "total_seniority_plus": round(total_seniority, 2),
            "total_overtime": round(total_overtime, 2),
            "total_gross": round(total_gross, 2),
            "total_ss_deduction": round(total_ss, 2),
            "total_irpf_deduction": round(total_irpf, 2),
            "total_net": round(total_net, 2),
            "total_employees": total_payrolls,
        },
        "employees": employee_costs,
        "page": page,
        "limit": limit,
        "total_employees": total_payrolls,
    }
