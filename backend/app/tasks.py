"""
TalentUP Fichaje — FastAPI BackgroundTasks helpers.

These functions are designed to run outside the request cycle. They create their
own DB session, perform the heavy work, and close the session when done.
"""
import logging
import os
import io
from datetime import date
from typing import Optional

from fastapi import BackgroundTasks
from fastapi.responses import StreamingResponse

from app.database import async_session_factory

logger = logging.getLogger(__name__)


def _iso_date(value: Optional[date]) -> Optional[str]:
    return value.isoformat() if value else None


async def run_payroll_close(
    tenant_id: str,
    user_id: str,
    month: int,
    year: int,
):
    """Background task: calculate and close payroll for a month."""
    from datetime import date as dt_date, datetime, time, timedelta, timezone
    from sqlalchemy import select

    from app.models.payroll import Payroll
    from app.models.employee import Employee
    from app.models.clock_in import ClockIn
    from app.models.shift import Shift
    from app.models.schedule import Schedule
    from app.models.overtime import Overtime
    from app.audit import log_action

    async with async_session_factory() as db:
        try:
            result = await db.execute(
                select(Employee).where(
                    Employee.tenant_id == tenant_id,
                    Employee.is_active == True,
                )
            )
            employees = result.scalars().all()

            start_date = dt_date(year, month, 1)
            if month == 12:
                end_date = dt_date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = dt_date(year, month + 1, 1) - timedelta(days=1)

            day_start = datetime.combine(start_date, time.min, tzinfo=timezone.utc)
            day_end = datetime.combine(end_date, time.max, tzinfo=timezone.utc)

            result = await db.execute(
                select(ClockIn).where(
                    ClockIn.tenant_id == tenant_id,
                    ClockIn.timestamp >= day_start,
                    ClockIn.timestamp <= day_end,
                    ClockIn.is_cancelled == False,
                ).order_by(ClockIn.employee_id, ClockIn.timestamp)
            )
            all_clock_ins = result.scalars().all()

            result = await db.execute(
                select(Overtime).where(
                    Overtime.tenant_id == tenant_id,
                    Overtime.date >= start_date,
                    Overtime.date <= end_date,
                )
            )
            all_overtime = result.scalars().all()

            result = await db.execute(
                select(Schedule).where(
                    Schedule.tenant_id == tenant_id,
                    Schedule.date >= start_date,
                    Schedule.date <= end_date,
                )
            )
            all_schedules = result.scalars().all()

            shift_result = await db.execute(select(Shift).where(Shift.tenant_id == tenant_id))
            all_shifts = {s.id: s for s in shift_result.scalars().all()}

            from collections import defaultdict

            clock_by_emp = defaultdict(list)
            for ci in all_clock_ins:
                clock_by_emp[ci.employee_id].append(ci)

            ot_by_emp = defaultdict(list)
            for o in all_overtime:
                ot_by_emp[o.employee_id].append(o)

            sched_by_emp = defaultdict(list)
            for s in all_schedules:
                sched_by_emp[s.employee_id].append(s)

            month_names = [
                "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
            ]
            period_label = f"{month_names[month]} {year}"
            created = []

            for emp in employees:
                existing = await db.execute(
                    select(Payroll).where(
                        Payroll.tenant_id == tenant_id,
                        Payroll.employee_id == emp.id,
                        Payroll.year == year,
                        Payroll.month == month,
                    )
                )
                if existing.scalar_one_or_none():
                    continue

                emp_clock = clock_by_emp.get(emp.id, [])
                emp_ot = ot_by_emp.get(emp.id, [])
                emp_sched = sched_by_emp.get(emp.id, [])

                total_worked_min = 0
                total_scheduled_min = 0
                holiday_min = 0
                late_min = 0
                early_leave_min = 0
                worked_days = set()

                daily_clock = defaultdict(list)
                for ci in emp_clock:
                    daily_clock[ci.timestamp.date()].append(ci)

                for day_date, day_clock in daily_clock.items():
                    first_in = None
                    last_out = None
                    for ci in day_clock:
                        if ci.type == "in":
                            if first_in is None or ci.timestamp < first_in:
                                first_in = ci.timestamp
                        elif ci.type == "out":
                            if last_out is None or ci.timestamp > last_out:
                                last_out = ci.timestamp

                    if first_in and last_out:
                        worked_seconds = (last_out - first_in).total_seconds()
                        total_worked_min += int(worked_seconds / 60)
                        worked_days.add(day_date)

                for s in emp_sched:
                    shift = all_shifts.get(s.shift_id)
                    if shift:
                        start_m = shift.start_time.hour * 60 + shift.start_time.minute
                        end_m = shift.end_time.hour * 60 + shift.end_time.minute
                        if end_m <= start_m:
                            end_m += 24 * 60
                        total_scheduled_min += end_m - start_m

                ot_structural = sum(o.total_minutes for o in emp_ot if o.overtime_type == "structural") / 60
                ot_fm = sum(o.total_minutes for o in emp_ot if o.overtime_type == "force_majeure") / 60
                ot_total = sum(o.total_minutes for o in emp_ot) / 60
                ot_amount = sum(float(o.overtime_amount or 0) for o in emp_ot)

                weekly_hours = float(emp.horas_semanales or 40)
                monthly_hours = weekly_hours * 4.33
                base_salary = float(emp.base_cotizacion or 0)
                coste_hora = float(emp.coste_hora or 0)
                if coste_hora > 0 and base_salary == 0:
                    base_salary = coste_hora * monthly_hours

                worked_hours = total_worked_min / 60
                scheduled_hours = total_scheduled_min / 60
                gross = base_salary + ot_amount
                ss_ded = gross * 0.0635
                irpf = gross * 0.12
                net = gross - ss_ded - irpf

                payroll = Payroll(
                    tenant_id=tenant_id,
                    employee_id=emp.id,
                    year=year,
                    month=month,
                    period_label=period_label,
                    contract_type=emp.tipo_contrato,
                    professional_category=emp.categoria_profesional,
                    work_day_type=emp.tipo_jornada,
                    weekly_hours=emp.horas_semanales,
                    scheduled_hours=round(scheduled_hours, 2),
                    worked_hours=round(worked_hours, 2),
                    worked_days=len(worked_days),
                    absent_days=0,
                    holiday_hours=round(holiday_min / 60, 2),
                    overtime_structural=round(ot_structural, 2),
                    overtime_force_majeure=round(ot_fm, 2),
                    overtime_total=round(ot_total, 2),
                    overtime_amount=round(ot_amount, 2),
                    late_minutes=late_min,
                    early_leave_minutes=early_leave_min,
                    base_salary=round(base_salary, 2),
                    gross_total=round(gross, 2),
                    ss_deduction=round(ss_ded, 2),
                    irpf_deduction=round(irpf, 2),
                    net_total=round(net, 2),
                    status="calculated",
                )
                db.add(payroll)
                created.append(payroll)

            await db.flush()
            await log_action(
                db,
                tenant_id=tenant_id,
                user_id=user_id,
                action="close_payroll",
                entity_type="payroll",
                new_value={"month": month, "year": year, "employees": len(created)},
            )
            await db.commit()
            logger.info(
                "payroll_closed tenant=%s month=%s year=%s employees=%s",
                tenant_id, month, year, len(created)
            )
        except Exception as exc:
            await db.rollback()
            logger.exception("payroll_close_failed tenant=%s month=%s year=%s: %s", tenant_id, month, year, exc)
            raise


async def run_report_export(
    format: str,
    tenant_id: Optional[str],
    date_from: str,
    date_to: str,
    employee_id: Optional[str],
    user_id: str,
    report_data: list[dict],
    job_id: str,
):
    """
    Background task: generate the PDF/Excel report and write it to a file so it
    can later be downloaded using the job_id.
    """
    from pathlib import Path

    exports_dir = Path(os.environ.get("EXPORTS_DIR", "exports"))
    exports_dir.mkdir(parents=True, exist_ok=True)

    try:
        if format == "pdf":
            filename = f"fichajes_{date_from}_{date_to}_{job_id}.pdf"
            response = _generate_pdf_sync(report_data, date_from, date_to)
        else:
            filename = f"fichajes_{date_from}_{date_to}_{job_id}.xlsx"
            response = _generate_excel_sync(report_data, date_from, date_to)

        body = b"".join([chunk async for chunk in response.body_iterator])
        file_path = exports_dir / filename
        file_path.write_bytes(body)

        logger.info(
            "report_export_done job_id=%s format=%s path=%s tenant=%s user=%s",
            job_id, format, file_path, tenant_id, user_id,
        )
    except Exception as exc:
        logger.exception("report_export_failed job_id=%s: %s", job_id, exc)
        raise


def _generate_pdf_sync(data, date_from, date_to):
    """Generate PDF report using reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        title=f"Informe de Fichajes {date_from} a {date_to}",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=18,
        textColor=colors.HexColor("#FF6B35"),
        spaceAfter=20,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.gray,
        spaceAfter=20,
    )

    elements = []
    elements.append(Paragraph("TalentUP Fichaje — Informe de Registro", title_style))
    elements.append(Paragraph(f"Período: {date_from} a {date_to}", subtitle_style))
    elements.append(Spacer(1, 10))

    for emp in data:
        elements.append(Paragraph(
            f"<b>{emp['name']}</b>  |  DNI: {emp['dni']}  |  Total: {emp['total_hours']}h",
            styles["Heading2"],
        ))
        elements.append(Spacer(1, 5))

        if not emp["entries"]:
            elements.append(Paragraph("Sin fichajes en este período", styles["Normal"]))
        else:
            table_data = [["Fecha", "Entrada", "Salida", "Horas"]]
            for entry in emp["entries"]:
                table_data.append([
                    entry["date"],
                    entry["in"],
                    entry["out"],
                    f"{entry['hours']}h",
                ])
            # Total row
            table_data.append(["", "", "TOTAL:", f"{emp['total_hours']}h"])

            col_widths = [100, 80, 80, 80]
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF6B35")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -2), 0.5, colors.grey),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF3EB")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F8F9FA")]),
            ]))
            elements.append(table)

        elements.append(Spacer(1, 15))

    # Footer with legal notice
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        "Documento generado por TalentUP Fichaje. "
        "Registro de jornada laboral conforme al RD-ley 8/2019 art. 34.9 ET. "
        "Conservación mínima: 4 años.",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7, textColor=colors.grey),
    ))

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=fichajes_{date_from}_{date_to}.pdf",
        },
    )


def _generate_excel_sync(data, date_from, date_to):
    """Generate Excel report using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Fichajes"

    # Header
    ws.merge_cells("A1:D1")
    ws["A1"] = f"TalentUP Fichaje — Informe {date_from} a {date_to}"
    ws["A1"].font = Font(size=14, bold=True, color="FF6B35")
    ws["A1"].alignment = Alignment(horizontal="center")

    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row = 3
    for emp in data:
        # Employee header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value=f"{emp['name']} — Total: {emp['total_hours']}h")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        # Table header
        for col, h in enumerate(["Fecha", "Entrada", "Salida", "Horas"], 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        row += 1

        # Data rows
        for entry in emp["entries"]:
            ws.cell(row=row, column=1, value=entry["date"]).border = thin_border
            ws.cell(row=row, column=2, value=entry["in"]).border = thin_border
            ws.cell(row=row, column=3, value=entry["out"]).border = thin_border
            ws.cell(row=row, column=4, value=entry["hours"]).border = thin_border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
            row += 1

        # Total row
        ws.cell(row=row, column=3, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4, value=emp["total_hours"]).font = Font(bold=True)
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
        row += 2  # blank row between employees

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=fichajes_{date_from}_{date_to}.xlsx",
        },
    )


# Keep the synchronous generators available for in-process / backwards-compatible use.


def _generate_pdf(data, date_from, date_to):
    return _generate_pdf_sync(data, date_from, date_to)


def _generate_excel(data, date_from, date_to):
    return _generate_excel_sync(data, date_from, date_to)


async def run_incident_detection(
    tenant_id: str,
    user_id: str,
    target_date: Optional[date] = None,
):
    """Background task: run incident detection for a tenant and date."""
    from app.incidents import detect_incidents

    if target_date is None:
        target_date = date.today()

    async with async_session_factory() as db:
        try:
            await detect_incidents(db, tenant_id, target_date)
            await db.commit()
            logger.info("incident_detection tenant=%s date=%s user=%s", tenant_id, target_date, user_id)
        except Exception as exc:
            await db.rollback()
            logger.exception("incident_detection_failed tenant=%s date=%s: %s", tenant_id, target_date, exc)
            raise


def schedule_payroll_close(
    background_tasks: BackgroundTasks,
    tenant_id: str,
    user_id: str,
    month: int,
    year: int,
):
    background_tasks.add_task(run_payroll_close, tenant_id, user_id, month, year)


def schedule_report_export(
    background_tasks: BackgroundTasks,
    format: str,
    tenant_id: Optional[str],
    date_from: str,
    date_to: str,
    employee_id: Optional[str],
    user_id: str,
    report_data: list[dict],
    job_id: str,
):
    background_tasks.add_task(
        run_report_export,
        format,
        tenant_id,
        date_from,
        date_to,
        employee_id,
        user_id,
        report_data,
        job_id,
    )


def schedule_incident_detection(
    background_tasks: BackgroundTasks,
    tenant_id: str,
    user_id: str,
    target_date: Optional[date] = None,
):
    background_tasks.add_task(run_incident_detection, tenant_id, user_id, target_date)
